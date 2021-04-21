#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sat Apr 10 00:47:14 2021

@author: z
"""

import openpyxl
import pandas as pd
import os
import datetime
import matplotlib.pyplot as plt

plt.close("all")

def tradeMa(ma,excelname,sheetname):
    
    wb = openpyxl.load_workbook(str(excelname)+".xlsx")
    ws = wb[str(sheetname)]
    
    dataset = []
    i = 1
    while(ws.cell(row=i, column=1).value!=None):
        row = []
        j = 1
        while(ws.cell(row=i, column=j).value!=None):
            row.append(ws.cell(row=i, column=j).value)
            j = j +1
        dataset.append(row)
        i = i + 1
    pd.DataFrame(dataset)
    
    # trade =[[0 for x in range(9)] for y in range(len(dataset))]
    trade = []
    for i in range(0, len(dataset)):
        row = []
        for j in range(0, 9):
            row.append(0)
        trade.append(row)
    trade[0][0] = "date"
    trade[0][1] = "buy"
    trade[0][2] = "hold"
    trade[0][3] = "sell"
    trade[0][4] = "number"
    trade[0][5] = "value"
    trade[0][6] = "period_return"
    trade[0][7] = "closeprice"
    trade[0][8] = "ma"

    for i in range(1, len(trade)):
        trade[i][0] = dataset[i][0]
        trade[i][7] = float(dataset[i][5])
    
    for i in range(ma, len(trade)):
        temp = 0
        for j in range(0, ma):
            temp = temp + trade[i-j][7]
        trade[i][8] = temp / ma
    
    for i in range(ma, len(trade)):
        if(trade[i][7] > trade[i][8])and(trade[i-1][2]==0):
            trade[i][1] = 1
            trade[i][2] = 1
            trade[i][3] = 0
        elif(trade[i][7] < trade[i][8])and(trade[i-1][2]==1):
            trade[i][1] = 0
            trade[i][2] = 0
            trade[i][3] = 1
        elif(trade[i-1][2]>0):
            trade[i][1] = 0
            trade[i][2] = 1
            trade[i][3] = 0
    
    initial = 10000
    for i in range(ma, len(trade)):
        if(trade[i][1] == 1)and(trade[i][2]==1)and(trade[i][3] == 0):
            trade[i][4] = initial/trade[i][7]
            trade[i][5] = trade[i][4] * trade[i][7]
            trade[i][6] = round((trade[i][4] * trade[i][7] - initial)/initial, 4)
        elif(trade[i][1] == 0)and(trade[i][2]==0)and(trade[i][3] == 1):
            trade[i][4] = 0
            trade[i][5] = trade[i-1][4] * trade[i][7]
            trade[i][6] = round((trade[i-1][4] * trade[i][7] - initial)/initial, 4)
        elif(trade[i][1] == 0)and(trade[i][2]==1)and(trade[i][3] == 0):
            trade[i][4] = trade[i-1][4]
            trade[i][5] = trade[i][4] * trade[i][7]
            trade[i][6] = round((trade[i][4] * trade[i][7] - initial)/initial, 4)
    #print(pd.DataFrame(trade))
    
    for i in range(0, len(trade)):
        trade[i].append(0)
        trade[i].append(0)
    trade[0][9]  = "holding_days"
    trade[0][10] = "Annualized Return"
    
    start = 0
    end = 0
    for i in range(1, len(trade)):
        if(trade[i][1]==1):
            trade[i][9] = 1
            start = datetime.datetime.strptime(str(trade[i][0]), "%Y-%m-%d %H:%M:%S").date()
        elif((trade[i][2] > 0) or (trade[i][3] == 1))and(trade[i][1] == 0):
            end = datetime.datetime.strptime(str(trade[i][0]), "%Y-%m-%d %H:%M:%S").date()
            delta = end - start
            trade[i][9] = delta.days+1
    
    for i in range(1, len(trade)):
        if(trade[i][1]==1)or(trade[i][2]==1)or(trade[i][3]==1):
            trade[i][10] = (1+trade[i][6])**(365/trade[i][9])-1
    
    #print(pd.DataFrame(trade))
            
    annual_return = 1
    trading_counts = 0
    finalResult=[]
    for i in range(1, len(trade)):
        if(trade[i][3]==1):
            annual_return = annual_return * (1+trade[i][6])
            trading_counts = trading_counts + 1
    #print(annual_return)
    #print(trading_counts)
    final_return = annual_return**(1/trading_counts)-1
    #print(trading_counts)
    finalResult=[ma,final_return,annual_return,trading_counts]
    return finalResult

excelname= "AAPL"
sheetname = "AAPL"
result=[]
for i in range(0, 60):
    row = []
    for j in range(0, 4):
        row.append(0)
    result.append(row)

for i in range(2,60):
    tradeMa(i,excelname,sheetname)
    result[i][0:4]=tradeMa(i,excelname,sheetname)[0:4]

df = pd.DataFrame(result, columns=['ma','final_return','annual_return','trading_counts'])
print(df)
print(df.loc[df['final_return'].idxmax()])












